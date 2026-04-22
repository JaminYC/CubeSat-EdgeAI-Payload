"""
Anade una hoja nueva 'INTISATv1+Payload' al Excel del pinout, replicando
exactamente el formato de INTISATv1 pero agregando una columna PAYLOAD
que documenta que pines usa la payload de microscopia (RPi 5).

Uso:
    python add_payload_pinout.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy
from pathlib import Path

XLSX = Path(__file__).parent / "Documentos de Referencia" / "INTISAT_PC-104_pinout.xlsx"

# Colores del original (BGR/ARGB Excel)
COLORS_ORIG = {
    "OBC-COMS":   "FFFFF2CC",  # cream (CAN bus)
    "OBC-UHF":    "FFD9EAD3",  # light green (UART)
    "OBC-EPS":    "FFFCE5CD",  # light peach (no se vio en data, asumido)
    "EPS-UHF":    "FFCFE2F3",  # light cyan
    "BUSES":      "FFC9DAF8",  # light blue (SPI/I2C)
    "3V3":        "FFEAD1DC",  # light pink
    "5V":         "FFD9D2E9",  # light purple
    "VBUS":       "FFFFFFFF",  # white (NC mostly)
    "GND":        "FFB27D0D",  # dark mustard
}

# Colores nuevos para nuestra payload
COLORS_PAYLOAD = {
    "PAYLOAD_PWR":   "FFFFD966",  # amarillo dorado: lineas de potencia consumidas
    "PAYLOAD_DATA":  "FF93C47D",  # verde fuerte: buses de datos consumidos (I2C/UART/SPI)
    "PAYLOAD_GPIO":  "FFEA9999",  # rojo claro: GPIOs de control/handshake
    "PAYLOAD_GND":   "FFB7B7B7",  # gris: GND compartido
}

# Mapeo de pines que la payload USA. Formato: (header, pin) -> (descripcion, color_key)
# Basado en Plan_CubeSat_RPi.md
PAYLOAD_USAGE = {
    # ── H1 ──
    ("H1", 16): ("RPi GPIO23 ← OBC (signal: payload busy)",  "PAYLOAD_GPIO"),
    ("H1", 17): ("RPi GPIO24 ← OBC (signal: payload error)", "PAYLOAD_GPIO"),
    ("H1", 21): ("RPi GPIO3 (I2C SCL slave addr 0x42)",      "PAYLOAD_DATA"),
    ("H1", 23): ("RPi GPIO2 (I2C SDA slave addr 0x42)",      "PAYLOAD_DATA"),
    ("H1", 29): ("GND comun RPi/sensor/OLED",                 "PAYLOAD_GND"),
    ("H1", 30): ("GND comun RPi/sensor/OLED",                 "PAYLOAD_GND"),
    ("H1", 31): ("GND comun RPi/sensor/OLED",                 "PAYLOAD_GND"),
    ("H1", 32): ("GND comun RPi/sensor/OLED",                 "PAYLOAD_GND"),
    ("H1", 33): ("GND comun RPi/sensor/OLED",                 "PAYLOAD_GND"),
    ("H1", 34): ("GND comun RPi/sensor/OLED",                 "PAYLOAD_GND"),
    ("H1", 45): ("VIN +5V RPi 5 (entrada principal)",         "PAYLOAD_PWR"),
    ("H1", 46): ("VIN +5V RPi 5 (entrada principal)",         "PAYLOAD_PWR"),
    # ── H2 ──
    ("H2", 3):  ("RPi UART0 RX (debug/comandos largos)",      "PAYLOAD_DATA"),
    ("H2", 4):  ("RPi UART0 TX (debug/comandos largos)",      "PAYLOAD_DATA"),
    ("H2", 25): ("VIN +5V sensor + OLED (3A)",                "PAYLOAD_PWR"),
    ("H2", 26): ("VIN +5V sensor + OLED (3A)",                "PAYLOAD_PWR"),
    ("H2", 29): ("DGND sensor/OLED",                          "PAYLOAD_GND"),
    ("H2", 30): ("DGND sensor/OLED",                          "PAYLOAD_GND"),
    ("H2", 31): ("DGND sensor/OLED",                          "PAYLOAD_GND"),
    ("H2", 32): ("DGND sensor/OLED",                          "PAYLOAD_GND"),
    ("H2", 35): ("3V3 OLED logica (si OLED requiere 3.3V)",   "PAYLOAD_PWR"),
    ("H2", 36): ("3V3 OLED logica (si OLED requiere 3.3V)",   "PAYLOAD_PWR"),
    ("H2", 39): ("RPi GPIO17 ← CPU_WD_1 (sync watchdog OBC)", "PAYLOAD_GPIO"),
    ("H2", 40): ("RPi GPIO27 ← CPU_MODE (detecta SAFE)",      "PAYLOAD_GPIO"),
}


def copy_cell_style(src, dst):
    """Copia formato (font, fill, border, alignment, etc) de una celda a otra."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def main():
    wb = openpyxl.load_workbook(XLSX)
    src_name = "INTISATv1"
    dst_name = "INTISATv1+Payload"

    # Si existe la hoja, eliminar para regenerar limpia
    if dst_name in wb.sheetnames:
        del wb[dst_name]

    # Copiar la hoja original (preserva formato)
    src = wb[src_name]
    dst = wb.copy_worksheet(src)
    dst.title = dst_name

    # Anadir leyenda de colores nuevos (debajo de la leyenda existente, fila 13-16 vacias)
    legend_start_row = 13
    bold_font = Font(bold=True)
    dst.cell(row=legend_start_row, column=2, value="LEYENDA PAYLOAD").font = bold_font
    legend_payload_rows = {
        "PAYLOAD_PWR":  "Alimentacion consumida por la payload (RPi/sensor/OLED)",
        "PAYLOAD_DATA": "Buses de datos consumidos (I2C esclavo / UART debug)",
        "PAYLOAD_GPIO": "GPIOs de control y handshake (busy/error/watchdog/safe)",
        "PAYLOAD_GND":  "GND compartido con la payload",
    }
    # Insertar despues de la fila 12 (GND legend); usar filas justo debajo
    # Como el original ya termina la leyenda en R12, usamos R13 en adelante.
    # Para no romper indices de las tablas H1/H2 (que empiezan en R15), usamos filas pequenas.
    # Mejor: agregamos al lado en columna H/I.
    legend_col = 8  # H
    dst.cell(row=3, column=legend_col, value="LEYENDA PAYLOAD").font = bold_font
    r = 4
    for key, desc in legend_payload_rows.items():
        cell_color = dst.cell(row=r, column=legend_col)
        cell_color.fill = PatternFill(start_color=COLORS_PAYLOAD[key],
                                       end_color=COLORS_PAYLOAD[key],
                                       fill_type="solid")
        cell_color.value = key
        dst.cell(row=r, column=legend_col + 1, value=desc)
        r += 1

    dst.column_dimensions['H'].width = 18
    dst.column_dimensions['I'].width = 65

    # Anadir titulo "PAYLOAD" en columna G de los headers H1 (R16) y H2 (R71)
    header_font = Font(bold=True)
    for header_row in (16, 71):
        cell = dst.cell(row=header_row, column=7, value="PAYLOAD")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Resaltar el encabezado H1/H2 (R15, R70) con la columna G tambien gris
    gray_fill = PatternFill(start_color="FF999999", end_color="FF999999", fill_type="solid")
    for hdr_row in (15, 70):
        for c in (3, 4, 5, 6, 7):
            cell = dst.cell(row=hdr_row, column=c)
            if cell.value is None:
                cell.fill = gray_fill

    # Mapeo de fila → (header, pin) para localizar donde escribir
    # H1: rows 17-68 = pins 1-52
    # H2: rows 72-123 = pins 1-52
    def row_for(header, pin):
        if header == "H1":
            return 16 + pin   # pin 1 → row 17
        elif header == "H2":
            return 71 + pin   # pin 1 → row 72
        return None

    # Rellenar columna G con los usos de la payload
    payload_col = 7
    for (header, pin), (desc, color_key) in PAYLOAD_USAGE.items():
        row = row_for(header, pin)
        if row is None:
            continue
        cell = dst.cell(row=row, column=payload_col, value=desc)
        cell.fill = PatternFill(start_color=COLORS_PAYLOAD[color_key],
                                 end_color=COLORS_PAYLOAD[color_key],
                                 fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Pines que NO usa la payload: poner "NC" igual que el resto, sin color
    for header in ("H1", "H2"):
        for pin in range(1, 53):
            if (header, pin) in PAYLOAD_USAGE:
                continue
            row = row_for(header, pin)
            cell = dst.cell(row=row, column=payload_col)
            if cell.value is None:
                cell.value = "NC"
                cell.alignment = Alignment(horizontal="center")

    # Ancho de la columna PAYLOAD
    dst.column_dimensions['G'].width = 50

    wb.save(XLSX)
    print(f"OK -> nueva hoja '{dst_name}' agregada a {XLSX.name}")


if __name__ == "__main__":
    main()
