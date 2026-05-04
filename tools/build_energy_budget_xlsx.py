"""
Genera el spreadsheet de presupuesto energetico del payload.

Hojas:
  1. Componentes — potencia idle/activo/pico de cada componente.
  2. Fases       — duracion y componentes activos por fase del pipeline.
  3. Por_scan    — energia total por scan (formulas auto-calculadas).
  4. Por_orbita  — presupuesto a nivel orbita.
  5. Mediciones  — celdas para llenar con valores reales del INA219.

Salida:
    Documentos de Referencia/Energy_Budget_Microscopia.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Documentos de Referencia" / "Energy_Budget_Microscopia.xlsx"

# Estilos
TITLE_FILL = PatternFill("solid", fgColor="1A365D")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2C5282")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FEF5E7")
RESULT_FILL = PatternFill("solid", fgColor="C6F6D5")
TOTAL_FILL = PatternFill("solid", fgColor="EBF8FF")
GRAY = PatternFill("solid", fgColor="F7FAFC")

THIN = Side(border_style="thin", color="CBD5E0")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def title(ws, row, col, text, span=4):
    ws.cell(row=row, column=col, value=text).font = TITLE_FONT
    ws.cell(row=row, column=col).fill = TITLE_FILL
    ws.cell(row=row, column=col).alignment = CENTER
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def data_row(ws, row, values, fill=None, start_col=1, formats=None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.alignment = CENTER if isinstance(v, (int, float)) else LEFT
        c.border = BORDER
        if fill is not None:
            c.fill = fill
        if formats and i < len(formats) and formats[i]:
            c.number_format = formats[i]


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────
wb = Workbook()

# ─── Hoja 1: Componentes ─────────────────────────────────────────────
ws = wb.active
ws.title = "Componentes"

title(ws, 1, 1, "Potencia por componente (datasheet)", span=6)
header_row(ws, 3, ["Componente", "P idle (W)", "P activo (W)",
                    "P pico (W)", "Cantidad", "Fuente"])

componentes = [
    ("Raspberry Pi 5 (4 GB)",       3.00, 5.50, 8.00, 1, "RPi Foundation"),
    ("OV5647 sensor",                0.05, 0.20, 0.30, 1, "OmniVision DS"),
    ("OLED SSD1351 (50% pixels)",    0.00, 0.25, 0.80, 1, "Solomon Systech"),
    ("TXS0108E level shifter",       0.005, 0.010, 0.020, 1, "TI"),
    ("MCP2515 CAN transceiver*",     0.040, 0.080, 0.100, 1, "Microchip (opcional)"),
    ("INA219 sensor I2C*",           0.005, 0.005, 0.005, 1, "TI (opcional)"),
]
for i, (name, idle, act, peak, qty, src) in enumerate(componentes, start=4):
    data_row(ws, i,
             [name, idle, act, peak, qty, src],
             fill=GRAY if i % 2 else None,
             formats=[None, "0.000", "0.000", "0.000", "0", None])

# Fila de totales con formulas
row_total = 4 + len(componentes)
data_row(ws, row_total,
         ["TOTAL (todos activos)",
          f"=SUMPRODUCT(B4:B{row_total-1},E4:E{row_total-1})",
          f"=SUMPRODUCT(C4:C{row_total-1},E4:E{row_total-1})",
          f"=SUMPRODUCT(D4:D{row_total-1},E4:E{row_total-1})",
          "—", "—"],
         fill=TOTAL_FILL,
         formats=[None, "0.00 \"W\"", "0.00 \"W\"", "0.00 \"W\"", None, None])
ws.cell(row=row_total, column=1).font = Font(bold=True)

ws.cell(row=row_total + 2, column=1, value="* Componentes opcionales").font = Font(italic=True, color="718096")

autosize(ws, [32, 14, 14, 14, 10, 24])
ws.row_dimensions[1].height = 24

# ─── Hoja 2: Fases ───────────────────────────────────────────────────
ws = wb.create_sheet("Fases")
title(ws, 1, 1, "Duracion y potencia por fase del pipeline", span=6)
header_row(ws, 3, ["Fase", "Duracion (s)", "P media estimada (W)",
                    "Componentes activos", "Energia E = P·t (J)", "Notas"])

fases = [
    ("IDLE pre-scan",    5,   3.05,  "RPi solo",                   "espera comando"),
    ("CAPTURING",        8,   5.95,  "RPi + OLED + sensor",        "25 angulos OLED"),
    ("PROCESSING",      30,   7.50,  "RPi (CPU 100%)",             "StarDist + CLAHE"),
    ("DOWNLINK",        10,   4.00,  "RPi + I2C",                  "thumbnail 100 KB"),
    ("IDLE post",        5,   3.05,  "RPi solo",                   "espera proximo cmd"),
]
for i, (fase, t, p, comp, notas) in enumerate(fases, start=4):
    data_row(ws, i,
             [fase, t, p, comp, f"=B{i}*C{i}", notas],
             fill=GRAY if i % 2 else None,
             formats=[None, "0", "0.00", None, "0.0 \"J\"", None])

row_total = 4 + len(fases)
data_row(ws, row_total,
         ["TOTAL POR SCAN",
          f"=SUM(B4:B{row_total-1})",
          f"=SUMPRODUCT(B4:B{row_total-1},C4:C{row_total-1})/SUM(B4:B{row_total-1})",
          "—",
          f"=SUM(E4:E{row_total-1})", "—"],
         fill=TOTAL_FILL,
         formats=[None, "0 \"s\"", "0.00 \"W\" \"prom\"", None,
                   "0.0 \"J\"", None])
ws.cell(row=row_total, column=1).font = Font(bold=True)

autosize(ws, [22, 14, 22, 28, 18, 28])
ws.row_dimensions[1].height = 24

# ─── Hoja 3: Por scan (calculo) ──────────────────────────────────────
ws = wb.create_sheet("Por_scan")
title(ws, 1, 1, "Resumen energetico por scan", span=4)

ws.cell(row=3, column=1, value="Magnitud").font = Font(bold=True, color="FFFFFF")
ws.cell(row=3, column=1).fill = HEADER_FILL
ws.cell(row=3, column=1).alignment = CENTER
ws.cell(row=3, column=1).border = BORDER
ws.cell(row=3, column=2, value="Valor").font = Font(bold=True, color="FFFFFF")
ws.cell(row=3, column=2).fill = HEADER_FILL
ws.cell(row=3, column=2).alignment = CENTER
ws.cell(row=3, column=2).border = BORDER

magnitudes = [
    ("Voltaje del rail",                "5.0",          "V"),
    ("Energia total por scan (J)",      "=Fases!E9",    "J"),
    ("Energia total por scan (Wh)",     "=B5/3600",     "Wh"),
    ("Carga (mAh) por scan",            "=(B5/B4)/3.6", "mAh"),
    ("Duracion total del scan (s)",     "=Fases!B9",    "s"),
    ("Potencia media",                  "=B5/B8",       "W"),
]
for i, (label, val, unit) in enumerate(magnitudes, start=4):
    fill_label = GRAY if i % 2 else None
    ws.cell(row=i, column=1, value=label).fill = fill_label or PatternFill()
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=1).alignment = LEFT
    ws.cell(row=i, column=2, value=val).fill = RESULT_FILL
    ws.cell(row=i, column=2).border = BORDER
    ws.cell(row=i, column=2).alignment = CENTER
    if "J" in unit:    ws.cell(row=i, column=2).number_format = "0.0 \"J\""
    elif "Wh" in unit: ws.cell(row=i, column=2).number_format = "0.0000 \"Wh\""
    elif "mAh" in unit: ws.cell(row=i, column=2).number_format = "0.00 \"mAh\""
    elif "W" in unit:  ws.cell(row=i, column=2).number_format = "0.00 \"W\""
    elif "V" in unit:  ws.cell(row=i, column=2).number_format = "0.00 \"V\""
    elif "s" in unit:  ws.cell(row=i, column=2).number_format = "0 \"s\""

autosize(ws, [32, 16])
ws.row_dimensions[1].height = 24

# ─── Hoja 4: Por orbita ──────────────────────────────────────────────
ws = wb.create_sheet("Por_orbita")
title(ws, 1, 1, "Presupuesto energetico por orbita", span=4)

# Inputs
ws.cell(row=3, column=1, value="ENTRADAS (modificar)").font = Font(bold=True)
ws.cell(row=3, column=1).fill = HEADER_FILL
ws.cell(row=3, column=1).font = Font(bold=True, color="FFFFFF")
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

inputs = [
    ("Duracion de la orbita (min)",     90,    "0 \"min\""),
    ("Numero de scans por orbita",      4,     "0"),
    ("Tiempo en SAFE / sleep (min)",    60,    "0 \"min\""),
    ("Potencia en SAFE (W)",            1.0,   "0.00 \"W\""),
    ("Tension del rail (V)",            5.0,   "0.00 \"V\""),
]
for i, (label, val, fmt) in enumerate(inputs, start=4):
    ws.cell(row=i, column=1, value=label).alignment = LEFT
    ws.cell(row=i, column=1).border = BORDER
    c = ws.cell(row=i, column=2, value=val)
    c.fill = INPUT_FILL
    c.border = BORDER
    c.number_format = fmt
    c.alignment = CENTER

# Outputs
row = 11
ws.cell(row=row, column=1, value="SALIDAS (calculadas)").font = Font(bold=True, color="FFFFFF")
ws.cell(row=row, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

outputs = [
    ("Energia activa por orbita (J)",          "=Por_scan!B5*B5",                       "0.0 \"J\""),
    ("Energia en SAFE por orbita (J)",         "=B6*60*B7",                             "0.0 \"J\""),
    ("Energia total por orbita (J)",           "=B12+B13",                              "0.0 \"J\""),
    ("Energia total por orbita (Wh)",          "=B14/3600",                             "0.000 \"Wh\""),
    ("Carga total por orbita (mAh)",           "=(B14/B8)/3.6",                         "0.0 \"mAh\""),
    ("Potencia promedio orbital (W)",          "=B14/(B4*60)",                          "0.00 \"W\""),
    ("Bateria sugerida (margen 50%, mAh)",     "=B16*1.5",                              "0 \"mAh\""),
]
for i, (label, val, fmt) in enumerate(outputs, start=12):
    ws.cell(row=i, column=1, value=label).alignment = LEFT
    ws.cell(row=i, column=1).border = BORDER
    c = ws.cell(row=i, column=2, value=val)
    c.fill = RESULT_FILL
    c.border = BORDER
    c.number_format = fmt
    c.alignment = CENTER

autosize(ws, [38, 16])
ws.row_dimensions[1].height = 24

# ─── Hoja 5: Mediciones reales (template) ────────────────────────────
ws = wb.create_sheet("Mediciones_reales")
title(ws, 1, 1, "Mediciones experimentales (INA219)", span=6)

ws.cell(row=3, column=1, value="Pegar aqui los valores reales medidos con tools/power_profiler.py para validar las estimaciones.").font = Font(italic=True, color="718096")
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

header_row(ws, 5, ["Fase", "P media medida (W)", "P pico medida (W)",
                    "Duracion medida (s)", "E medida (J)",
                    "Diferencia vs estimado (%)"])

fases_med = ["IDLE pre-scan", "CAPTURING", "PROCESSING", "DOWNLINK", "IDLE post"]
for i, fase in enumerate(fases_med, start=6):
    ws.cell(row=i, column=1, value=fase).alignment = LEFT
    ws.cell(row=i, column=1).border = BORDER
    for col in range(2, 6):
        c = ws.cell(row=i, column=col, value="")
        c.fill = INPUT_FILL
        c.border = BORDER
    ws.cell(row=i, column=5, value=f"=B{i}*D{i}").fill = RESULT_FILL
    ws.cell(row=i, column=5).number_format = "0.0 \"J\""
    ws.cell(row=i, column=6, value=f"=IFERROR(((B{i}-Fases!C{i-2})/Fases!C{i-2})*100,\"\")").fill = RESULT_FILL
    ws.cell(row=i, column=6).number_format = "0.0 \"%\""
    ws.cell(row=i, column=6).border = BORDER

autosize(ws, [22, 18, 18, 18, 16, 22])
ws.row_dimensions[1].height = 24

# Guardar
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUT))
print(f"OK: {OUT}")
